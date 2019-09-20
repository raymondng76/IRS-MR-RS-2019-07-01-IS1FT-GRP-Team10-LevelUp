import openpyxl
from openpyxl import load_workbook
import nltk
from nltk.probability import *
from nltk.corpus import stopwords  # 3
import re  # 4
from collections import Counter # 5
import datetime

# 1.load file

wb = load_workbook('competency_extraction_demo.xlsx')
ws = wb["sheet1"]
threshold = 5
save = 'competency_extraction_result.xlsx'

print(f"======>>>>>{datetime.datetime.now()} Start")
row_list = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6, min_col=6, values_only=True):
    row_list += row
text_description = str(row_list).upper()

# 2.tokenize

text_0 = nltk.word_tokenize(text_description)

# 3.stopwords

stop = set(stopwords.words("english"))
text_1 = []
for token in text_0:
    if token not in stop:
        text_1.append(token)

# 4.delete punctuation mark

punctuation = re.compile(r'[\\\n\$#-.?!,":;()|0-9|`/]')
text_2 = []
for token in text_1:
    if len(token) <= 1:
        word = punctuation.sub("",token)
        text_2.append(word)
        while '' in text_2:
            text_2.remove('')
    else:
        # avoid delete punctuation mark within the word (to reserve 'C++', 'word2vec' )
        text_2.append(token)
print(text_2)

# 5. combine n-gram word list
nfdist_1 = FreqDist(list(nltk.ngrams(text_2, 1)))
key_1 = []
for key in dict(Counter(nfdist_1)).keys():
    key_01 = ''
    for key_n in list(key):
        if key_n != list(key)[-1]:
            key_01 += key_n + " "
        else :
            key_01 += key_n
    key_1.append(key_01)
newdict_1 = dict(zip(list(key_1), list(dict(Counter(nfdist_1)).values())))

nfdist_2 = FreqDist(list(nltk.ngrams(text_2, 2)))
key_2 = []
for key in dict(Counter(nfdist_2)).keys():
    key_02 = ''
    for key_n in list(key):
        if key_n != list(key)[-1]:
            key_02 += key_n + " "
        else :
            key_02 += key_n
    key_2.append(key_02)
newdict_2 = dict(zip(list(key_2), list(dict(Counter(nfdist_2)).values())))

nfdist_3 = FreqDist( list(nltk.ngrams(text_2, 3)))
key_3 = []
for key in dict(Counter(nfdist_3)).keys():
    key_03 = ''
    for key_n in list(key):
        if key_n != list(key)[-1]:
            key_03 += key_n + " "
        else :
            key_03 += key_n
    key_3.append(key_03)
newdict_3 = dict(zip(list(key_3), list(dict(Counter(nfdist_3)).values())))
print(f"======>>>>>{datetime.datetime.now()} generated the name and frequency of all ngrams")


def divider(excel_terminology,n):
    """1.divide terminology word list into different groups by the number of words；

    *Please make sure the abbreviation has been split from the terminology before using this function.
    for example : convert ['natural language processing (NLP)'] into ['natural language processing'],['NLP']

    :param excel_terminology: file name of the terminology dictionary
    :param n: n word list
    """
    from openpyxl import load_workbook


    wb = load_workbook(excel_terminology)
    ws = wb.active
    term_list_1 = []
    term_list_2 = []
    term_list_3 = []
    space_num = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, min_col=1, values_only=True):
        str_row = "".join(row)
        for space in str_row:
            if space == " ":
                space_num += 1
        if space_num == 0:
            term_list_1.append(str_row)
        elif space_num == 1:
            term_list_2.append(str_row)
        elif space_num == 2:
            term_list_3.append(str_row)
        space_num = 0  # reset space caculator
    if n == 1:
        return term_list_1
    if n == 2:
        return term_list_2
    if n == 3:
        return term_list_3

# 7. match job description content with encyclopedia(bag-of-words),high frequency words are job competency we want.

key_1 = []
key_2 = []
key_3 = []
value_1 = []
value_2 = []
value_3 = []

for key in newdict_1.keys():
    if threshold < newdict_1.get(key):
        if key in divider('terminology_modified_1.xlsx', 1):
            key_1.append(key)
            value_01= newdict_1.get(key)
            value_1.append(value_01)
print(f"======>>>>>{datetime.datetime.now()} 1-word list match complete")
for key in newdict_2.keys():
    if threshold < newdict_2.get(key):
        if key in divider('terminology_modified_2.xlsx', 2):
            key_2.append(key)
            value_02 = newdict_2.get(key)
            value_2.append(value_02)
print(f"======>>>>>{datetime.datetime.now()} 2-word list match complete")
for key in newdict_3.keys():
    if threshold < newdict_3.get(key):
        if key in divider('terminology_modified_3.xlsx', 3):
            key_3.append(key)
            value_03 = newdict_3.get(key)
            value_3.append(value_03)
print(f"======>>>>>{datetime.datetime.now()} 3-word list match complete")

print(f"======>>>>>{datetime.datetime.now()} generated the name and frequency of skill ngrams")
newdict_1 = dict(zip(key_1, value_1))
newdict_2 = dict(zip(key_2, value_2))
newdict_3 = dict(zip(key_3, value_3))
print(newdict_1)
print(newdict_2)
print(newdict_3)

# 7. export excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'skill'
data_excel = []
for each in newdict_1:
    data_excel.append([each, newdict_1[each]])
for each in newdict_2:
    data_excel.append([each, newdict_2[each]])
for each in newdict_3:
    data_excel.append([each, newdict_3[each]])
for each in data_excel:
    ws.append(each)
wb.save(save)
print( f'======>>>>>{datetime.datetime.now()} excel has been created for skill ngrams' )








